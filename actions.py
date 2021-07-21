# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"

from typing import Any, Text, Dict, List

from rasa_sdk import Action, Tracker
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet
from rasa_sdk.forms import FormAction
from rasa_sdk.events import FollowupAction
from rasa_sdk.events import AllSlotsReset
from weather import get_weather
from rasa_sdk import Tracker, FormValidationAction
from rasa_sdk.types import DomainDict
import requests
import openpyxl
import re
import datetime
import sqlite3
import database
#from rasa.core.trackers import DialogueStateTracker

# class ActionHelloWorld(Action):
#
#     def name(self) -> Text:
#         return "action_hello_world"
#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []

class ActionFacilitySearch(Action):

   def name(self) -> Text:
        return "action_facility_search"

   def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        location = tracker.get_slot("location")
        dispatcher.utter_message(text="Facility Search Action!")

        return [AllSlotsReset()]

#checking 



class ActionWeather(Action):

    def name(self) -> Text:
        return "action_weather"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        city = tracker.get_slot('location')
        weather_data = get_weather(city)
        temp = weather_data['main']['temp']
        min_temp = weather_data['main']['temp_min']
        max_temp = weather_data['main']['temp_max']
        desc = weather_data['weather'][0]['description']
        humidity = weather_data['main']['humidity']
        response = "Weather in {} is {}, Temperature is {} degree Celsius with minimum temperature {} degree Celsius and maximum temperature {} degree Celsius, Humidity is {}%".format(city, desc, temp, min_temp, max_temp, humidity)
        dispatcher.utter_message(response)

        return [SlotSet('location', city)]        


class ActionRestart(Action):

  def name(self) -> Text:
      return "action_restart"

  async def run(
      self, dispatcher, tracker: Tracker, domain: Dict[Text, Any]
  ) -> List[Dict[Text, Any]]:

      # custom behavior

      return []

# class ActionPoliceStation(Action):

#     def name(self) -> Text:
#          return "action_police_station"

#     def run(self, dispatcher: CollectingDispatcher,
#              tracker: Tracker,
#              domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
            
#          location = tracker.get_slot("location")           

        
#          message = "police station contact number : 100 "
#          final_message = message + "in " + location  
         
#         # print(final_message) 
#          dispatcher.utter_message(text= final_message)

             

#          return [AllSlotsReset()]




class PoliceNumbers(Action):
    def name(self) -> Text:
        return "police_number_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["city"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
               

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitpn(Action):
    def name(self) -> Text:
        return "action_dbpn"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
        city = tracker.get_slot("city")    

        final_message = "NATIONAL EMERGENCY NUMBER:112\nPOLICE:100\nFIRE:101\nAMBULANCE:102\nDisaster Management Services:108\nWomen Helpline:1091\nDomestic Abuse:181\nAir Ambulance:9540161344\nAids Helpline:1097\n"        
        dispatcher.utter_message(text= final_message) 
  

        return [AllSlotsReset()]        





                                 
       
 #-------------------------------------------------------------------------            

#other fir form
class ReportFIROtherForm(Action):
    def name(self) -> Text:
        return "user_details_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","problem"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

           # All slots are filled.
        return [SlotSet("requested_slot", None)]


class Actionsubmittingother(Action):
    def name(self) -> Text:
        return "action_checking"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name= tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        problem = tracker.get_slot("problem")



        database.otherfir(name,mobile_no, aadhar_no,email, dob, problem)     


        dispatcher.utter_message(text = "thank you for the information, " + tracker.get_slot("name"))
        return [AllSlotsReset()]        

# class ActionSubmit(Action):
#     def name(self) -> Text:
#         return "action_submit_other"

#     def run(
#         self,
#         dispatcher,
#         tracker: Tracker,
#         domain: "DomainDict",
#     ) -> List[Dict[Text, Any]]:
       



#         #database.otherfir(tracker.get_slot("name"),tracker.get_slot("mobile_no"), tracker.get_slot("aadhar_no"), tracker.get_slot("email"), tracker.get_slot("dob"), tracker.get_slot("problem"))     


#         dispatcher.utter_message(text = tracker.get_slot("name")+ "hi")
#         return [AllSlotsReset()]

 #-------------------------------------------------------------------------
# vehicle theft form
class ReportFIROvehicletheftForm(Action):
    def name(self) -> Text:
        return "theft_of_vehicle_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","vehicle","number_plate","colour","last_seen","anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class Actiondbvt(Action):
    def name(self) -> Text:
        return "action_databasevt"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name= tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        vehicle = tracker.get_slot("vehicle")
        number_plate= tracker.get_slot("number_plate")
        colour = tracker.get_slot("colour")
        last_seen = tracker.get_slot("last_seen")
        anything_else = tracker.get_slot("anything_else")



        database.vehicletheft(name ,mobile_no, aadhar_no,email, dob, vehicle,number_plate,colour, last_seen, anything_else)     


        dispatcher.utter_message(text = "thank you for the information, " + tracker.get_slot("name"))
        return [AllSlotsReset()]             


 #-------------------------------------------------------------------------        
#theft form

class ReportFIRgoodstheftForm(Action):
    def name(self) -> Text:
        return "theft_of_goods_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","goods_stolen","date_time","last_seen","anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class Actiondbgt(Action):
    def name(self) -> Text:
        return "action_databasegt"



    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name = tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        goods_stolen = tracker.get_slot("goods_stolen")
        date_time= tracker.get_slot("date_time")        
        last_seen = tracker.get_slot("last_seen")
        anything_else = tracker.get_slot("anything_else")

        database.goods_theft(name ,mobile_no , aadhar_no , email , dob, goods_stolen, date_time, last_seen, anything_else)


       

        dispatcher.utter_message(text = "thank you for the information ," + tracker.get_slot("name"))
        return [AllSlotsReset()] 
 #-------------------------------------------------------------------------

class ReportFIRMissingPersonForm(Action):
    def name(self) -> Text:
        return "missing_person_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","name_of_missing_person","age_of_mm","sex_of_mm","clothes_of_mm","last_seen","anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitmpdb(Action):
    def name(self) -> Text:
        return "action_mpdb"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name = tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        name_of_missing_person = tracker.get_slot("name_of_missing_person")
        age_of_mm= tracker.get_slot("age_of_mm")     
        sex_of_mm= tracker.get_slot("sex_of_mm")
        clothes_of_mm = tracker.get_slot("clothes_of_mm")   
        last_seen = tracker.get_slot("last_seen")
        anything_else = tracker.get_slot("anything_else")

        database.missingperson(name, mobile_no,aadhar_no,email, dob,name_of_missing_person,age_of_mm,sex_of_mm,clothes_of_mm,last_seen,anything_else)
               
        dispatcher.utter_message(text = "thank you for the information ," + tracker.get_slot("name"))
        return [AllSlotsReset()]   
 #-------------------------------------------------------------------------                                                    

class ReportLostandFoundForm(Action):
    def name(self) -> Text:
        return "lost_and_found_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","lost_or_found","item_lost_or_found","location_of_item","anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitlf(Action):
    def name(self) -> Text:
        return "action_dblf"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
        name = tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        lost_or_found = tracker.get_slot("lost_or_found")
        item_lost_or_found= tracker.get_slot("item_lost_or_found")
        location_of_item= tracker.get_slot("location_of_item")
        anything_else = tracker.get_slot("anything_else")

        database.lostandfound(name, mobile_no,aadhar_no,email, dob,lost_or_found,item_lost_or_found,location_of_item,anything_else)

        dispatcher.utter_message(text = "thank you for the information ," + tracker.get_slot("name"))          

       
        return [AllSlotsReset()]  
#-------------------------------------------------------------------------        

class ReportAssaultForm(Action):
    def name(self) -> Text:
        return "assault_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","email", "dob","complaint_against","assault_description","anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitas(Action):
    def name(self) -> Text:
        return "action_dbas"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name = tracker.get_slot("name")
        mobile_no =tracker.get_slot("mobile_no")
        aadhar_no =tracker.get_slot("aadhar_no")
        email =tracker.get_slot("email") 
        dob =  tracker.get_slot("dob")
        complaint_against = tracker.get_slot("complaint_against")
        assault_description= tracker.get_slot("assault_description")
        anything_else = tracker.get_slot("anything_else")

        database.assault(name, mobile_no,aadhar_no,email, dob, complaint_against,assault_description,anything_else)
           

        dispatcher.utter_message(text = "thank you for the information ," + tracker.get_slot("name"))
        return [AllSlotsReset()]  
#-------------------------------------------------------------------------        


class ReportCivicGrievanceForm(Action):
    def name(self) -> Text:
        return "civic_grievance_form"

    def run(
        self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name", "mobile_no","aadhar_no","civic_grievance", "location","state", "landmark", "anything_else"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]
                

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitcg(Action):
    def name(self) -> Text:
        return "action_dbcg"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        name = tracker.get_slot("name")
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")
        civic_grievance = tracker.get_slot("civic_grievance")
        location = tracker.get_slot("location")
        state = tracker.get_slot("state")
        landmark = tracker.get_slot("landmark")
        anything_else = tracker.get_slot("anything_else")

        database.civic_grievance(name, mobile_no,aadhar_no,civic_grievance, location,state, landmark, anything_else)         


        dispatcher.utter_message(text = "thank you for the information, " + tracker.get_slot("name"))
        return [AllSlotsReset()]

#--------------------------------------------------------------------------------------
class MedicalFacilityForm(Action):
    def name(self) -> Text:
        return "medical_facility_form"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["name_of_ailment", "name", "age", "dob", "blood_group", "sex", "marital_status", "mobile_no", "aadhar_no", "current_address", "city", "pincode"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

#-----------------------------------------------------------------------------------

class MedicalFacilityFormSubmit(Action):
    def name(self) -> Text:
        return "action_dbmf"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:

        name_of_ailment = tracker.get_slot("name_of_ailment")
        name = tracker.get_slot("name")
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob")
        blood_group = tracker.get_slot("blood_group")
        sex = tracker.get_slot("sex")
        marital_status = tracker.get_slot("marital_status")
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode")

        database.hospital(name_of_ailment,name,age ,dob,blood_group, sex,marital_status,mobile_no,aadhar_no,current_address ,city, pincode)
        dispatcher.utter_message(text = "thank you for the information, " + tracker.get_slot("name"))
        return [AllSlotsReset()]


#----------------------------------------------------------------------------------
class DrivingLicenseForm1(Action):
    def name(self) -> Text:
        return "driving_license_form1"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitDL1(Action):
    def name(self) -> Text:
        return "action_dbdlgearless"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        learners_licence_no = tracker.get_slot("learners_license_no")
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.dl_mc_gearless(learners_licence_no, name ,age , dob, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]

#---------------------------------------------------------------------------------------------
class DrivingLicenseForm2(Action):
    def name(self) -> Text:
        return "driving_license_form2"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitDL2(Action):
    def name(self) -> Text:
        return "action_dbdlgear"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        learners_licence_no = tracker.get_slot("learners_license_no")
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.dl_mc_gear(learners_licence_no, name ,age , dob, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]

#---------------------------------------------------------------------------------------------
class DrivingLicenseForm3(Action):
    def name(self) -> Text:
        return "driving_license_form3"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitDL3(Action):
    def name(self) -> Text:
        return "action_dbdlchv"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        learners_licence_no = tracker.get_slot("learners_license_no")
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.dl_chv(learners_licence_no, name ,age , dob, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]
#------------------------------------------------------------------------------------------
class LearnersLicenseForm1(Action):
    def name(self) -> Text:
        return "learners_license_form1"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "pob", "citizenship_type", "education", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitLL1(Action):
    def name(self) -> Text:
        return "action_dbllgearless"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        pob = tracker.get_slot("pob")
        citizenship_type= tracker.get_slot("citizenship_type")
        education= tracker.get_slot("education")
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.ll_mc_gearless( name ,age , dob, pob,citizenship_type, education, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]

#---------------------------------------------------------------------------------------------
class LearnersLicenseForm2(Action):
    def name(self) -> Text:
        return "learners_license_form2"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "pob", "citizenship_type", "education", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitLL2(Action):
    def name(self) -> Text:
        return "action_dbllgear"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        pob = tracker.get_slot("pob")
        citizenship_type= tracker.get_slot("citizenship_type")
        education= tracker.get_slot("education")
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.ll_mc_gear(name ,age , dob, pob,citizenship_type, education, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]
#---------------------------------------------------------------------------------------------
class LearnersLicenseForm3(Action):
    def name(self) -> Text:
        return "learners_license_form3"

    def run(
            self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict
    ) -> List[Dict[Text, Any]]:
        required_slots = ["learner_license_no", "name", "age", "dob", "pob", "citizenship_type", "education", "mobile_no", "aadhar_no", "current_address", "city", "pincode", "date_for_appointment", "slot_for_appointment"]

        for slot_name in required_slots:
            if tracker.slots.get(slot_name) is None:
                # The slot is not filled yet. Request the user to fill this slot next.
                return [SlotSet("requested_slot", slot_name)]

        # All slots are filled.
        return [SlotSet("requested_slot", None)]

class ActionSubmitLL3(Action):
    def name(self) -> Text:
        return "action_dbllchv"

    def run(
        self,
        dispatcher,
        tracker: Tracker,
        domain: "DomainDict",
    ) -> List[Dict[Text, Any]]:
         
        name = tracker.get_slot("name") 
        age = tracker.get_slot("age")
        dob = tracker.get_slot("dob") 
        pob = tracker.get_slot("pob")
        citizenship_type= tracker.get_slot("citizenship_type")
        education= tracker.get_slot("education")
        mobile_no = tracker.get_slot("mobile_no")
        aadhar_no = tracker.get_slot("aadhar_no")  
        current_address = tracker.get_slot("current_address")
        city = tracker.get_slot("city")
        pincode = tracker.get_slot("pincode") 
        date_for_appointment = tracker.get_slot("date_for_appointment")
        slot_for_appointment = tracker.get_slot("slot_for_appointment")

        

        database.ll_chv(name ,age , dob, pob,citizenship_type, education, mobile_no , aadhar_no , current_address ,city,pincode ,date_for_appointment ,slot_for_appointment)

        return[AllSlotsReset()]        
#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateUserDetailsForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_user_details_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_problem(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `problem` value."""

        # If the problem is super short, it might be wrong.
        print(f"problem = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"problem": None}
        else:
            return {"problem": slot_value}        

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}
#--------------------------------------------------------------------------
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        #aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (tracker.get_slot('aadhar_no') == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, tracker.get_slot('aadhar_no'))):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------


#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateTheftOfVehicleForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_theft_of_vehicle_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_vehicle(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `vehicle` value."""

        # If the problem is super short, it might be wrong.
        print(f"vehicle = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"vehicle": None}
        else:
            return {"vehicle": slot_value}     

#--------------------------------------------------------------------------
    def validate_colour(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `colour` value."""

        # If the problem is super short, it might be wrong.
        print(f"vehicle = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's very short. I'm assuming you mis-spelled")
            return {"colour": None}
        else:
            return {"colour": slot_value}                    

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}
#--------------------------------------------------------------------------

    def validate_last_seen(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last seen` value."""

        # If the name is super short, it might be wrong.
        print(f"last seen = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"last_seen": None}
        else:
            return {"last_seen": slot_value}

#--------------------------------------------------------------------------           
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

#--------------------------------------------------------------------------
    def validate_number_plate(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `number plate` value."""

        print(f"vehicle identification number given = {slot_value}")
        number_plate = tracker.get_slot('number_plate')
        regex = ("^([A-Z|a-z]{2}\s{1}\d{2}\s{1}[A-Z|a-z]{1,2}\s{1}\d{1,4})?([A-Z|a-z]{3}\s{1}\d{1,4})?$")

        p = re.compile(regex)

        if (number_plate == None):
            dispatcher.utter_message(text=f"Please enter a valid vehicle identification number.")
            return {"number_plate": None}

        if (re.search(p, number_plate)):
            return {"number_plate": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid vehicle identification number.")
            return {"number_plate": None}            
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------


#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateTheftOfGoodsForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_theft_of_goods_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_goods_stolen(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `goods_stolen` value."""

        # If the problem is super short, it might be wrong.
        print(f"goods_stolen = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"goods_stolen": None}
        else:
            return {"goods_stolen": slot_value}     
          

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}
#--------------------------------------------------------------------------

    def validate_last_seen(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last seen` value."""

        # If the name is super short, it might be wrong.
        print(f"last seen = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"last_seen": None}
        else:
            return {"last_seen": slot_value}

#--------------------------------------------------------------------------           
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
   
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------

#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateMissingPersonForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_missing_person_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_name_of_missing_person(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name of missing person` value."""

        # If the problem is super short, it might be wrong.
        print(f"name_of_missing_person = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name_of_missing_person": None}
        else:
            return {"name_of_missing_person": slot_value}     

#--------------------------------------------------------------------------
    def validate_colour_of_mm(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name of missing person` value."""

        # If the problem is super short, it might be wrong.
        print(f"colour of clothes = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's very vague. please provide more details")
            return {"colour_of_mm": None}
        else:
            return {"colour_of_mm": slot_value}                 
          

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}
#--------------------------------------------------------------------------

    def validate_last_seen(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last seen` value."""

        # If the name is super short, it might be wrong.
        print(f"last seen = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"last_seen": None}
        else:
            return {"last_seen": slot_value}

#--------------------------------------------------------------------------           
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
   
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------

# lost and found form

class ValidateLostAndFoundForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_lost_and_found_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_item_lost_or_found(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name of missing person` value."""

        # If the problem is super short, it might be wrong.
        print(f"item_lost_or_found = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"item_lost_or_found": None}
        else:
            return {"item_lost_or_found": slot_value}     

#--------------------------------------------------------------------------
    def validate_location_of_item(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name of missing person` value."""

        # If the problem is super short, it might be wrong.
        print(f"location_of_item = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's very vague. please provide more details")
            return {"location_of_item": None}
        else:
            return {"location_of_item": slot_value}                 
          

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

#--------------------------------------------------------------------------           
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
   
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------

#assault fir validation

#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateAssaultFirForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_assault_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------

    def validate_complaint_against(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"complaint_against = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"complaint_against": None}
        else:
            return {"complaint_against": slot_value}
#--------------------------------------------------------------------------
    def validate_assault_description(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `assault_description` value."""

        # If the problem is super short, it might be wrong.
        print(f"assault_description = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"assault_description": None}
        else:
            return {"assault_description": slot_value}        

#--------------------------------------------------------------------------
    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}
#--------------------------------------------------------------------------
    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}
#--------------------------------------------------------------------------
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------
# civic grievance validation

#-------------------------------------------------------------------------
#--------------------------------------------------------------------------
#--------------------------------------------------------------------------

class ValidateCivicGrievanceForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_civic_grievance_form"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"Name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}
#--------------------------------------------------------------------------
    def validate_grievance(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `problem` value."""

        # If the problem is super short, it might be wrong.
        print(f"grievance = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 5:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"grievance": None}
        else:
            return {"grievance": slot_value}        



#--------------------------------------------------------------------------
    def validate_landmark(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `landmark` value."""

        # If the problem is super short, it might be wrong.
        print(f"landmark = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 8:
            dispatcher.utter_message(text=f"That's a little vague. Please provide more details")
            return {"landmark": None}
        else:
            return {"landmark": slot_value}        



#--------------------------------------------------------------------------
    def validate_email(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"email given = {slot_value}")
        email = tracker.get_slot('email')
        regex = ("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+$")

        p = re.compile(regex)

        if (email == None):
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}

        if (re.search(p, email)):
            return {"email": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid email.")
            return {"email": None}
#--------------------------------------------------------------------------

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

#--------------------------------------------------------------------------


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhar_no number given = {slot_value}")
        aadhar_no = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhar_no == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhar_no)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit aadhar_no number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}
#--------------------------------------------------------------------------------


  



#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []





class ValidateMedicalFacilityForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_medical_facility_form"

    def validate_name_of_ailment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name_of_ailment` value."""

        # If the name is super short, it might be wrong.
        print(f"name of ailment given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short input. I'm assuming you mis-spelled.")
            return {"name_of_ailment": None}
        else:
            return {"name_of_ailment": slot_value}

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

   
    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `age` value."""

        if not slot_value.isnumeric():
            return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if int(present_year - input_year) != int(age)+1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_blood_group(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `blood_group` value."""
        print(f"blood group given = {slot_value}")
        regex = ("^(A|B|AB|O)[+-]$")
        p = re.compile(regex)
        if slot_value is None:
            dispatcher.utter_message(text=f"Please enter a valid blood group and in proper format.")
            return {"blood_group": None}
        if re.search(p, slot_value):
            return {"blood_group": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid blood group and in proper format.")
            return {"blood_group": None}

    def validate_sex(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `sex` value."""

        print(f"sex given = {slot_value}")
        if slot_value.title() != "Male" and slot_value.title() != "Female" and slot_value.title() != "Intersex":
            dispatcher.utter_message(text=f"Please enter a valid option from Male, Female or Intersex.")
            return {"sex": None}
        else:
            return {"sex": slot_value}

    def validate_marital_status(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `marital_status` value."""

        # If the name is super short, it might be wrong.
        print(f"marital status given = {slot_value}")
        if slot_value.title() != "Married" and slot_value.title() != "Single" and slot_value.title() != "Divorced" and slot_value.title() != "Widowed":
            dispatcher.utter_message(text=f"Please enter a valid option from Married, Single, Divorced or Widowed.")
            return {"marital_status": None}
        else:
            return {"marital_status": slot_value}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhaar)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot('city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

class ValidateDrivingLicenseForm1(FormValidationAction):

    def name(self) -> Text:
        return "validate_driving_license_form1"

    def validate_learners_license_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `learner_license_no` value."""

        driving_license_type = tracker.get_slot('driving_license_type')
        # If the name is super short, it might be wrong.
        print(f"Learners license number given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) != 9 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Please enter a valid 9 digit learners license number.")
            return {"learners_license_no": None}
        else:
            return {"learners_license_n0": slot_value}

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

 
    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `driving_license_type` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if not slot_value.isnumeric():
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
                return {"age": None}
            elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
                return {"age": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
                return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age)+1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}


    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhaar)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        final_url = f"https://api.postalpincode.in/pincode/" + slot_value
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot('city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateDrivingLicenseForm2(FormValidationAction):

    def name(self) -> Text:
        return "validate_driving_license_form2"

    def validate_learner_license_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `learner_license_no` value."""

        driving_license_type = tracker.get_slot('driving_license_type')
        # If the name is super short, it might be wrong.
        print(
            f"Learners license number given = {slot_value} length = {len(slot_value)} driving license type = {driving_license_type}")
        if len(slot_value) != 9 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Please enter a valid 9 digit learners license number.")
            return {"learner_license_no": None}
        else:
            return {"learner_license_no": slot_value}

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `driving_license_type` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if not slot_value.isnumeric():
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
                return {"age": None}
            elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
                return {"age": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
                return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhaar)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateDrivingLicenseForm3(FormValidationAction):

    def name(self) -> Text:
        return "validate_driving_license_form3"

    def validate_learner_license_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `learner_license_no` value."""

        driving_license_type = tracker.get_slot('driving_license_type')
        # If the name is super short, it might be wrong.
        print(
            f"Learners license number given = {slot_value} length = {len(slot_value)} driving license type = {driving_license_type}")
        if len(slot_value) != 9 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Please enter a valid 9 digit learners license number.")
            return {"learner_license_no": None}
        else:
            return {"learner_license_no": slot_value}

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `driving_license_type` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if not slot_value.isnumeric():
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
                return {"age": None}
            elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
                return {"age": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
                dispatcher.utter_message(
                    text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
                return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if (re.search(p, aadhaar)):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateLearnersLicenseForm1(FormValidationAction):

    def name(self) -> Text:
        return "validate_learners_license_form1"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}



    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
            return {"age": None}
        elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
            return {"age": None}
        elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
            return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            print(driving_license_type)
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_pob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pob` value."""

        print(f"pob given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        pob = tracker.get_slot('pob')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == pob.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"pob": None}
        else:
            return {"pob": slot_value}

    def validate_citizenship_type(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `citizenship type` value."""

        print(f"citizenship type given = {slot_value}")
        if slot_value != "birth" and slot_value != "descent" and slot_value != "registration" and slot_value != "naturalisation":
            dispatcher.utter_message(text=f"Please enter a valid citizenship type.")
            return {"citizenship_type": None}
        else:
            return {"citizenship_type": slot_value}

    def validate_education(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `education` value."""

        print(f"education = {slot_value}")
        if len(slot_value) < 2:
            dispatcher.utter_message(text=f"That is a very short educational qualification. I am assuming you mis-spelled.")
            return {"education": None}
        else:
            return {"education": slot_value}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if aadhaar is None:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if re.search(p, aadhaar):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6:
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateLearnersLicenseForm2(FormValidationAction):

    def name(self) -> Text:
        return "validate_learners_license_form2"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
            return {"age": None}
        elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
            return {"age": None}
        elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
            return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            print(driving_license_type)
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_pob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pob` value."""

        print(f"pob given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        pob = tracker.get_slot('pob')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == pob.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"pob": None}
        else:
            return {"pob": slot_value}

    def validate_citizenship_type(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `citizenship type` value."""

        print(f"citizenship type given = {slot_value}")
        if slot_value != "birth" and slot_value != "descent" and slot_value != "registration" and slot_value != "naturalisation":
            dispatcher.utter_message(text=f"Please enter a valid citizenship type.")
            return {"citizenship_type": None}
        else:
            return {"citizenship_type": slot_value}

    def validate_education(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `education` value."""

        print(f"education = {slot_value}")
        if len(slot_value) < 2:
            dispatcher.utter_message(
                text=f"That is a very short educational qualification. I am assuming you mis-spelled.")
            return {"education": None}
        else:
            return {"education": slot_value}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if aadhaar is None:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if re.search(p, aadhaar):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6:
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateLearnersLicenseForm3(FormValidationAction):

    def name(self) -> Text:
        return "validate_learners_license_form3"

    def validate_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        # If the name is super short, it might be wrong.
        print(f"name = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"name": None}
        else:
            return {"name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
            return {"age": None}
        elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
            return {"age": None}
        elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
            return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            print(driving_license_type)
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_pob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pob` value."""

        print(f"pob given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        pob = tracker.get_slot('pob')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == pob.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"pob": None}
        else:
            return {"pob": slot_value}

    def validate_citizenship_type(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `citizenship type` value."""

        print(f"citizenship type given = {slot_value}")
        if slot_value != "birth" and slot_value != "descent" and slot_value != "registration" and slot_value != "naturalisation":
            dispatcher.utter_message(text=f"Please enter a valid citizenship type.")
            return {"citizenship_type": None}
        else:
            return {"citizenship_type": slot_value}

    def validate_education(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `education` value."""

        print(f"education = {slot_value}")
        if len(slot_value) < 2:
            dispatcher.utter_message(
                text=f"That is a very short educational qualification. I am assuming you mis-spelled.")
            return {"education": None}
        else:
            return {"education": slot_value}

    def validate_mobile_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `mobile_no` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('mobile_no')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

        if (re.search(p, phone)):
            return {"mobile_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"mobile_no": None}

    def validate_aadhar_no(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhar_no` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhar_no')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if aadhaar is None:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

        if re.search(p, aadhaar):
            return {"aadhar_no": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhar_no": None}

    def validate_current_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `current_address` value."""

        # If the name is super short, it might be wrong.
        print(f"current_address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short current_address. I'm assuming you mis-spelled.")
            return {"current_address": None}
        else:
            return {"current_address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pincode(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pincode` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pincode = tracker.get_slot('pincode')
        final_url = f"https://api.postalpincode.in/pincode/" + pincode
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6:
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pincode": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pincode": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pincode": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}

class ActionResetAllSlots(Action):
    def name(self):
        return "action_reset_all_slots"

    def run(self, dispatcher, tracker, domain):
        return [AllSlotsReset()]


class ActionCoronaTracker(Action):
    def name(self) -> Text:
        return "action_corona_tracker"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        response = requests.get("https://api.covid19india.org/data.json").json()

        entities = tracker.latest_message['entities']
        print("Last Message Now ", entities)
        state = None

        for e in entities:
            if e['entity'] == "state":
                state = e['value']
        if state.lower() == "india":
            state = "Total"
        message = "Please enter correct state name"
        for data in response['statewise']:
            if data['state'] == state.title() or data['statecode'] == state.upper():
                message = """Covid19 status in {} is
                    Active: {}, Confirmed: {}, Recovered: {} On {}""".format(state.title(), data["active"], data["confirmed"], data["recovered"], data["lastupdatedtime"])

        dispatcher.utter_message(message)

        return []

