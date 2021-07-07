# This files contains your custom actions which can be used to run
# custom Python code.
#
# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions


# This is a simple example for a custom action which utters "Hello World!"
import datetime
from typing import Any, Text, Dict, List
from rasa_sdk import Action, Tracker, FormValidationAction
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet, AllSlotsReset
from weather import get_weather
from rasa_sdk.types import DomainDict
import requests
import openpyxl
import re

# class ActionHelloWorld(Action):
#
#     def name(self) -> Text:
#         return "action_hello_world"
#
#     def run(self, dispatcher: CollectingDispatcher,
#             tracker: Tracker,
#             domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
#
#         dispatcher.utter_message(text="Hello World!")
#
#         return []


class ActionFacilitySearch(Action):

    def name(self) -> Text:
        return "action_facility_search"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # entities = tracker.latest_message['entities']
        dispatcher.utter_message(text="Facility Search Action!")

        return []


class ActionWeather(Action):

    def name(self) -> Text:
        return "action_weather"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        location = tracker.get_slot('location')
        weather_data = get_weather(location)
        temp = weather_data['main']['temp']
        min_temp = weather_data['main']['temp_min']
        max_temp = weather_data['main']['temp_max']
        desc = weather_data['weather'][0]['description']
        humidity = weather_data['main']['humidity']
        response = "Weather in {} is {}, Temperature is {} degree Celsius with minimum temperature {} degree Celsius and maximum temperature {} degree Celsius, Humidity is {}%".format(location, desc, temp, min_temp, max_temp, humidity)
        dispatcher.utter_message(response)

        return [SlotSet('location', location)]


class ValidateMedicalFacilityForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_medical_facility_form"

    def validate_name_of_ailment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `name_of_ailment` value."""

        # If the name is super short, it might be wrong.
        print(f"name of ailment given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short input. I'm assuming you mis-spelled.")
            return {"name_of_ailment": None}
        else:
            return {"name_of_ailment": slot_value}

    def validate_first_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `first_name` value."""

        # If the name is super short, it might be wrong.
        print(f"First name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"first_name": None}
        else:
            return {"first_name": slot_value}

    def validate_middle_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `middle_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Middle name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"middle_name": None}
        else:
            return {"middle_name": slot_value}

    def validate_last_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Last name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"last_name": None}
        else:
            return {"last_name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `driving_license_type` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if not slot_value.isnumeric():
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
                return {"age": None}
            elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
                return {"age": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
                return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age)+1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_blood_group(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `blood_group` value."""

        print(f"blood group given = {slot_value}")
        regex = "/^(A|B|AB|O)[+-]$/i"
        p = re.compile(regex)
        if slot_value is None:
            dispatcher.utter_message(text=f"Please enter a valid blood group and in proper format.")
            return {"blood_group": None}

        if re.search(p, slot_value):
            return {"blood_group": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid blood group and in proper format.")
            return {"blood_group": None}

    def validate_sex(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `sex` value."""

        print(f"sex given = {slot_value}")
        if slot_value.title() != "Male" and slot_value.title() != "Female" and slot_value.title() != "Intersex":
            dispatcher.utter_message(text=f"Please enter a valid option from Male, Female or Intersex.")
            return {"sex": None}
        else:
            return {"sex": slot_value}

    def validate_marital_status(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `marital_status` value."""

        # If the name is super short, it might be wrong.
        print(f"marital status given = {slot_value}")
        if slot_value.title() != "Married" and slot_value.title() != "Single" and slot_value.title() != "Divorced" and slot_value.title() != "Widowed":
            dispatcher.utter_message(text=f"Please enter a valid option from Married, Single, Divorced or Widowed.")
            return {"marital_status": None}
        else:
            return {"marital_status": slot_value}

    def validate_phone_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `phone_number` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('phone_number')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}

        if (re.search(p, phone)):
            return {"phone_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}

    def validate_aadhaar_card_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhaar_card_number` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhaar_card_number')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

        if (re.search(p, aadhaar)):
            return {"aadhaar_card_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

    def validate_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `address` value."""

        # If the name is super short, it might be wrong.
        print(f"address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short address. I'm assuming you mis-spelled.")
            return {"address": None}
        else:
            return {"address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pin_code(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pin_code` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pin_code = tracker.get_slot('pin_code')
        final_url = f"https://api.postalpincode.in/pincode/" + pin_code
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pin_code": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot('city').title():
                return {"pin_code": slot_value}
            else:
                dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pin_code": None}

class ValidateDrivingLicenseForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_driving_license_form"

    def validate_learners_license_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `learners_license_number` value."""

        driving_license_type = tracker.get_slot('driving_license_type')
        # If the name is super short, it might be wrong.
        print(f"Learners license number given = {slot_value} length = {len(slot_value)} driving license type = {driving_license_type}")
        if len(slot_value) != 9 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Please enter a valid 9 digit learners license number.")
            return {"learners_license_number": None}
        else:
            return {"learners_license_number": slot_value}

    def validate_first_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `first_name` value."""

        # If the name is super short, it might be wrong.
        print(f"First name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"first_name": None}
        else:
            return {"first_name": slot_value}

    def validate_middle_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `middle_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Middle name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"middle_name": None}
        else:
            return {"middle_name": slot_value}

    def validate_last_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Last name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"last_name": None}
        else:
            return {"last_name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `driving_license_type` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if not slot_value.isnumeric():
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
                return {"age": None}
            elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
                return {"age": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
                dispatcher.utter_message(text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
                return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age)+1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_phone_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `phone_number` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('phone_number')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}

        if (re.search(p, phone)):
            return {"phone_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}


    def validate_aadhaar_card_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhaar_card_number` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhaar_card_number')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if (aadhaar == None):
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

        if (re.search(p, aadhaar)):
            return {"aadhaar_card_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

    def validate_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `address` value."""

        # If the name is super short, it might be wrong.
        print(f"address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short address. I'm assuming you mis-spelled.")
            return {"address": None}
        else:
            return {"address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `city` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pin_code(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pin_code` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pin_code = tracker.get_slot('pin_code')
        final_url = f"https://api.postalpincode.in/pincode/" + pin_code
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6 and not slot_value.isnumeric():
            dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pin_code": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot('city').title():
                return {"pin_code": slot_value}
            else:
                dispatcher.utter_message(text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pin_code": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ValidateLearnersLicenseForm(FormValidationAction):

    def name(self) -> Text:
        return "validate_learners_license_form"

    def validate_first_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `first_name` value."""

        # If the name is super short, it might be wrong.
        print(f"First name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"first_name": None}
        else:
            return {"first_name": slot_value}

    def validate_middle_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `middle_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Middle name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"middle_name": None}
        else:
            return {"middle_name": slot_value}

    def validate_last_name(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `last_name` value."""

        # If the name is super short, it might be wrong.
        print(f"Last name given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 2:
            dispatcher.utter_message(text=f"That's a very short name. I'm assuming you mis-spelled.")
            return {"last_name": None}
        else:
            return {"last_name": slot_value}

    def validate_age(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `first_name` value."""
        driving_license_type = tracker.get_slot('driving_license_type')

        print(f"age given = {slot_value} driving_license_type = {driving_license_type}")
        if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and int(slot_value) < 16:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 16 years of age.")
            return {"age": None}
        elif driving_license_type == "motorcycles with gear" and int(slot_value) < 18:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 18 years of age.")
            return {"age": None}
        elif driving_license_type == "commercial heavy vehicles or transport vehicles" and int(slot_value) < 20:
            dispatcher.utter_message(
                text=f"Seems like you have entered a wrong value. If you are eligible for {driving_license_type} and already have a learners license, then you must be above 20 years of age.")
            return {"age": None}
        else:
            return {"age": slot_value}

    def validate_dob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `dob` value."""

        print(f"dob given = {slot_value}")
        driving_license_type = tracker.get_slot('driving_license_type')
        today = datetime.datetime.now()
        age = tracker.get_slot('age')
        present_year = today.year
        date_str = tracker.get_slot('dob')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            input_year = datetime_obj.year
            print(datetime_obj.date())
            print(driving_license_type)
            if driving_license_type == "motorcycles without gear, upto 50 cc capacity" and present_year - input_year < 16:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 16, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "motorcycles with" and present_year - input_year < 18:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 18, the age for eligibility.")
                return {"dob": None}
            elif driving_license_type == "commercial heavy vehicles or transport vehicles" and present_year - input_year < 20:
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it makes your age less than 20, the age for eligibility.")
                return {"dob": None}
            elif int(present_year - input_year) != int(age) + 1 and int(present_year - input_year) != int(age):
                dispatcher.utter_message(
                    text="Your birth year doesn't seem to be correct as it doesn't match your age.")
                return {"dob": None}
            else:
                return {"dob": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered your date of birth in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"dob": None}

    def validate_pob(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pob` value."""

        print(f"pob given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        pob = tracker.get_slot('pob')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == pob.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"pob": None}
        else:
            return {"pob": slot_value}

    def validate_citizenship_type(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `citizenship type` value."""

        print(f"citizenship type given = {slot_value}")
        if slot_value != "birth" and slot_value != "descent" and slot_value != "registration" and slot_value != "naturalisation":
            dispatcher.utter_message(text=f"Please enter a valid citizenship type.")
            return {"citizenship_type": None}
        else:
            return {"citizenship_type": slot_value}

    def validate_education(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `education` value."""

        print(f"education = {slot_value}")
        if len(slot_value) < 2:
            dispatcher.utter_message(text=f"That is a very short educational qualification. I am assuming you mis-spelled.")
            return {"education": None}
        else:
            return {"education": slot_value}

    def validate_phone_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `phone_number` value."""

        print(f"phone number given = {slot_value}")
        phone = tracker.get_slot('phone_number')
        regex = ("^(?:(?:\+|0{0,2})91(\s*[\-]\s*)?|[0]?)?[789]\d{9}$")

        p = re.compile(regex)

        if (phone == None):
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}

        if (re.search(p, phone)):
            return {"phone_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid phone number.")
            return {"phone_number": None}

    def validate_aadhaar_card_number(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `aadhaar_card_number` value."""

        print(f"aadhaar number given = {slot_value}")
        aadhaar = tracker.get_slot('aadhaar_card_number')
        regex = ("^[2-9]{1}[0-9]{3}\\" +
                 "s[0-9]{4}\\s[0-9]{4}$")

        p = re.compile(regex)

        if aadhaar is None:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

        if re.search(p, aadhaar):
            return {"aadhaar_card_number": slot_value}
        else:
            dispatcher.utter_message(text=f"Please enter a valid 12-digit Aadhaar number in XXXX XXXX XXXX format.")
            return {"aadhaar_card_number": None}

    def validate_address(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `address` value."""

        # If the name is super short, it might be wrong.
        print(f"address given = {slot_value} length = {len(slot_value)}")
        if len(slot_value) <= 10:
            dispatcher.utter_message(text=f"That's a very short address. I'm assuming you mis-spelled.")
            return {"address": None}
        else:
            return {"address": slot_value}

    def validate_city(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `first_name` value."""

        print(f"city given = {slot_value}")
        file = "IndianCitiesDb.xlsx"
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        state_of_city = ""
        city = tracker.get_slot('city')
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=214, max_col=1):
            for cell in row:
                if cell.value == city.title():
                    state_of_city = ws.cell(row=cell.row, column=2).value

        if state_of_city == "":
            dispatcher.utter_message(text=f"Seems like you have mis-spelled. Please enter a correct city name.")
            return {"city": None}
        else:
            print(state_of_city)
            SlotSet("state_for_license", state_of_city)
            return {"city": slot_value}

    def validate_pin_code(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `pin_code` value."""

        print(f"pin code given = {slot_value} length = {len(slot_value)}")
        pin_code = tracker.get_slot('pin_code')
        final_url = f"https://api.postalpincode.in/pincode/" + pin_code
        pin_data = requests.get(final_url).json()
        if len(slot_value) != 6:
            dispatcher.utter_message(
                text=f"Seems like you have entered an incorrect pin code. A pin code must be of 6 digits.")
            return {"pin_code": None}
        else:
            if pin_data[0]['Status'] == 'Success' and pin_data[0]['PostOffice'][0]['District'] == tracker.get_slot(
                    'city').title():
                return {"pin_code": slot_value}
            else:
                dispatcher.utter_message(
                    text=f"Seems like you have entered an incorrect pin code or your pin code doesn't match your city. Please enter a valid pin code.")
                return {"pin_code": None}

    def validate_date_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `date_for_appointment` value."""

        print(f"appointment date given = {slot_value}")
        today = datetime.datetime.now()
        date_str = tracker.get_slot('date_for_appointment')
        format_str = '%d-%m-%Y'  # The format
        try:
            datetime_obj = datetime.datetime.strptime(date_str, format_str)
            print(datetime_obj.date())
            if datetime_obj.date() < today.date():
                dispatcher.utter_message(
                    text="Seems like you have entered an incorrect date. Please enter a valid date.")
                return {"date_for_appointment": None}
            elif datetime_obj.weekday() == 6:
                dispatcher.utter_message(text="Sorry, we are closed on Sunday, please choose another date.")
                return {"date_for_appointment": None}
            else:
                return {"date_for_appointment": slot_value}
        except ValueError:
            print("incorrect date format")
            dispatcher.utter_message(
                text="Seems like you have entered the date in the incorrect format or have entered an incorrect date. Please use DD-MM-YYYY format and enter a valid date.")
            return {"date_for_appointment": None}

    def validate_slot_for_appointment(
            self,
            slot_value: Any,
            dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: DomainDict,
    ) -> Dict[Text, Any]:
        """Validate `slot_for_appointment` value."""

        selected_slot = tracker.get_slot('slot_for_appointment')
        print(f"slot name given = {slot_value}")
        if selected_slot != "10.00 to 12.00" and selected_slot != "12.30 to 14.30" and selected_slot != "15.00 to 17.00":
            dispatcher.utter_message(text=f"Please enter a valid slot.")
            return {"slot_for_appointment": None}
        else:
            return {"slot_for_appointment": slot_value}


class ActionResetAllSlots(Action):
    def name(self):
        return "action_reset_all_slots"

    def run(self, dispatcher, tracker, domain):
        return [AllSlotsReset()]


class ActionCoronaTracker(Action):
    def name(self) -> Text:
        return "action_corona_tracker"

    def run(self, dispatcher: CollectingDispatcher,
            tracker: Tracker,
            domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:

        response = requests.get("https://api.covid19india.org/data.json").json()

        entities = tracker.latest_message['entities']
        print("Last Message Now ", entities)
        state = None

        for e in entities:
            if e['entity'] == "state":
                state = e['value']
        if state.lower() == "india":
            state = "Total"
        message = "Please enter correct state name"
        for data in response['statewise']:
            if data['state'] == state.title() or data['statecode'] == state.upper():
                message = """Covid19 status in {} is
Active: {}, Confirmed: {}, Recovered: {} On {}""".format(state.title(), data["active"], data["confirmed"], data["recovered"], data["lastupdatedtime"])

        dispatcher.utter_message(message)

        return []
